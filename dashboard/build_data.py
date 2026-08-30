# -*- coding: utf-8 -*-
"""
בונה את קובץ הנתונים של הדשבורד מתוך קובצי המקור.

קלט   : "עיבוד לפי ענף ורשות בני דורמבט.xlsx"  (למ"ס, עיבוד מרכז הידע)
        "dicAnaf4SfarotMaster 2.xlsx"          (מילון ענפי כלכלה, למ"ס)
פלט   : dashboard/data.json  +  הזרקה ל-index.html / artifact.html דרך build.py

הרצה  : python3 dashboard/build_data.py
"""
import json, os, sys
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC_DATA = os.path.join(ROOT, "עיבוד לפי ענף ורשות בני דורמבט.xlsx")
SRC_DIC  = os.path.join(ROOT, "dicAnaf4SfarotMaster 2.xlsx")
SRC_RASH = os.path.join(ROOT, "p_libud_24.xlsx")   # קובץ רשויות מקומיות, מהדורת 2024
SRC_2022 = os.path.join(ROOT, "עיבוד לפי ענף ורשות 2022.xlsx")
SRC_BTL  = os.path.join(ROOT, "שכר ממוצע לחודש עבודה של כלל העובדים, לפי מחוז ונפה, 2022-2016.xlsx")

# --- שמות תצוגה קצרים לקבוצות הענפים (54 קודים דו-ספרתיים ב-51 קבוצות) -------
# נגזרו ממילון ענפי הכלכלה של הלמ"ס; קוצרו כדי שייקראו היטב על גבי גרף.
ANAF_LABELS = {
    "1":              "גידולים צמחיים",
    "2+3+4":          "בעלי חיים, דיג וייעור",
    "5+6+7+8":        "כרייה וחציבה",
    "10":             "ייצור מוצרי מזון",
    "11+12":          "ייצור משקאות וטבק",
    "13+14+15":       "טקסטיל, הלבשה ועור",
    "16+17+18":       "עץ, נייר ודפוס",
    "19+20":          "נפט מזוקק וכימיקלים",
    "21":             "ייצור תרופות",
    "22":             "גומי ופלסטיק",
    "23+24+25":       "מינרלים אל-מתכתיים ומתכת",
    "26":             "מחשבים, אלקטרוניקה ואופטיקה",
    "27+28":          "ציוד חשמלי ומכונות",
    "29+30":          "כלי רכב וכלי תחבורה",
    "31+32+33+34":    "רהיטים, יהלומים וייצור אחר",
    "35+36+37+38+39": "חשמל, מים וטיפול בפסולת",
    "41":             "בניית מבנים ובניינים",
    "42":             "עבודות הנדסה אזרחית",
    "43":             "עבודות בנייה מיוחדות",
    "45":             "מסחר ותיקון כלי רכב",
    "46":             "מסחר סיטוני",
    "47":             "מסחר קמעוני",
    "49":             "הובלה יבשתית",
    "50+51+52+53":    "הובלה ימית ואווירית, אחסנה ודואר",
    "55":             "שירותי אירוח",
    "56":             "מסעדות ושירותי מזון",
    "58+59+60":       "הוצאה לאור, קולנוע ושידור",
    "61":             "שירותי תקשורת",
    "62+63":          "תכנות, ייעוץ ושירותי מידע",
    "64+65+66":       "שירותים פיננסיים וביטוח",
    "68":             "פעילויות בנדל\"ן",
    "69":             "שירותים משפטיים וחשבונאות",
    "70":             "ייעוץ ניהולי ומשרדים ראשיים",
    "71":             "אדריכלות, הנדסה ובדיקות טכניות",
    "72":             "מחקר מדעי ופיתוח",
    "73+74+75":       "פרסום ושירותים מקצועיים אחרים",
    "77+79":          "השכרה וסוכנויות נסיעות",
    "78":             "שירותי תעסוקה",
    "80":             "שמירה ואבטחה",
    "81":             "תחזוקת בניינים וגינון",
    "82":             "ניהול ותמיכה למשרדים ולעסקים",
    "83":             "מינהל מקומי",
    "84":             "מינהל ציבורי, ביטחון וביטוח לאומי",
    "85":             "חינוך",
    "86":             "שירותי בריאות",
    "87+88":          "רווחה, סעד ומגורים טיפוליים",
    "90+91+92":       "אמנות, תרבות ובידור",
    "93":             "ספורט, בילוי ופנאי",
    "94":             "ארגוני חברים",
    "95+96":          "תיקון ושירותים אישיים",
    "97+98+99":       "משקי בית כמעסיקים וארגונים חוץ-מדינתיים",
}

# --- קיבוצים (clusters) ------------------------------------------------------
# הבסיס: עמודת SemelSederAnafOt/ShemSederAnafOt במילון — סדרי הענפים של הלמ"ס,
# אותיות A–U, שהן רמת הקיבוץ הרשמית מעל הרמה הדו-ספרתית שבה מגיעים הנתונים.
# כל קבוצת ענפים בנתונים שויכה לסדר של הקודים שבה; סדרים קרובים אוחדו כדי
# להימנע מקיבוצים בני ענף אחד:
#   B (כרייה) + D (חשמל) + E (מים ופסולת)  →  "אנרגיה, מים, פסולת וכרייה"
#   K (פיננסים) + L (נדל"ן)                →  "פיננסים, ביטוח ונדל\"ן"
#   R (אמנות) + S (שירותים אחרים) + T + U  →  "אמנות, פנאי ושירותים אחרים"
# "היי-טק" אינו סדר אלא קיבוץ חוצה-סדרים, שנגזר מהעמודות high_tech_manufacturing
# ו-high_tech_services שבמילון (ענפים 21, 26, 61, 62+63, 72) — ולכן הוא חופף
# לקיבוצים "תעשייה וייצור", "מידע ותקשורת" ו"שירותים מקצועיים".
CLUSTERS = [
    ("agri",    "חקלאות, ייעור ודיג",        ["1", "2+3+4"]),
    ("infra",   "אנרגיה, מים, פסולת וכרייה",  ["5+6+7+8", "35+36+37+38+39"]),
    ("manuf",   "תעשייה וייצור",              ["10", "11+12", "13+14+15", "16+17+18", "19+20", "21", "22",
                                               "23+24+25", "26", "27+28", "29+30", "31+32+33+34"]),
    ("constr",  "בינוי",                      ["41", "42", "43"]),
    ("trade",   "מסחר",                       ["45", "46", "47"]),
    ("transp",  "תחבורה, אחסנה ודואר",         ["49", "50+51+52+53"]),
    ("hosp",    "אירוח ומזון",                ["55", "56"]),
    ("ict",     "מידע ותקשורת",               ["58+59+60", "61", "62+63"]),
    ("fin",     "פיננסים, ביטוח ונדל\"ן",      ["64+65+66", "68"]),
    ("prof",    "שירותים מקצועיים ומדעיים",    ["69", "70", "71", "72", "73+74+75"]),
    ("admin",   "שירותי ניהול ותמיכה",         ["77+79", "78", "80", "81", "82"]),
    ("public",  "מינהל מקומי וציבורי",         ["83", "84"]),
    ("edu",     "חינוך",                      ["85"]),
    ("health",  "בריאות, רווחה וסעד",          ["86", "87+88"]),
    ("other",   "אמנות, פנאי ושירותים אחרים",  ["90+91+92", "93", "94", "95+96", "97+98+99"]),
    ("hitech",  "היי-טק (תעשייה ושירותים)",    ["21", "26", "61", "62+63", "72"]),
]

NAFOT = {  # שיוך רשות לנפה — נגזר מסדר הגיליון ומאומת מול סכומי הנפות
    "נפת צפת": ["ראש פינה", "יסוד המעלה", "מטולה", "ג'ש (גוש חלב)", "טובא-זנגרייה",
                "חצור הגלילית", "קריית שמונה", "הגליל העליון", "מרום הגליל",
                "מבואות החרמון", "צפת"],
    "נפת גולן": ["בוקעאתא", "קצרין", "מג'דל שמס", "מסעדה", "ע'ג'ר", "עין קנייא", "גולן"],
}


def num(v):
    return round(v, 4) if isinstance(v, (int, float)) else None


def read_dictionary():
    """מחזיר מיפוי קוד דו-ספרתי → {שם, סדר ענפים}."""
    wb = openpyxl.load_workbook(SRC_DIC, data_only=True)
    ws = wb["רשימת הערכים במילון ענפי כלכלה"]
    rows = list(ws.iter_rows(values_only=True))
    ix = {str(h).strip(): i for i, h in enumerate(rows[0]) if h}
    out = {}
    for r in rows[1:]:
        code = r[ix["SemelAnaf2Sfarot"]]
        if code is None:
            continue
        code = str(code).strip()
        if code in out:
            continue
        out[code] = {
            "name": (r[ix["ShemAnaf2Sfarot"]] or "").strip(),
            "seder": (r[ix["ShemSederAnafOt"]] or "").strip(),
            "letter": (r[ix["SemelSederAnafOt"]] or "").strip(),
        }
    wb.close()
    return out


# הקובץ של 2022 מציג רק פילוח רשות × ענף, ללא שורת סיכום לרשות. הסכימה נעשית
# מרמת "סדר ענפים" (אותיות) ולא מרמת שתי הספרות: שתי הרמות מדווחות רק תאים שבהם
# 10 שכירים ומעלה, ולכן ברמה הדקה אובדים 5.2% מהשכירים לעומת 0.5% לכל היותר ברמת
# הסדר. הממוצע משוקלל לפי מספר השכירים — אותה שיטה שבה הלמ"ס עצמה גוזרת את שורות
# הנפה מנתוני הרשויות בקובץ 2024 (נבדק והתאים בדיוק).
SUPPRESSION_FLOOR = 10


def read_btl_trend():
    """
    סדרת מגמה של הביטוח הלאומי, 2016–2022, ברמת האשכול מול הארצי.

    הלוח מפורסם לפי מחוז ונפה. רמת האשכול נבנית מחיבור נפת צפת ונפת גולן —
    אותה גאוגרפיה בדיוק של 18 הרשויות (נבדק מול קובץ הלמ"ס: סכום השכירים
    ב-18 הרשויות זהה לסכום שתי הנפות).

    שים לב: בקובץ קיימת שורת "גליל מזרחי" מוכנה, אך השכר בה הוא ממוצע פשוט
    של שתי הנפות ולא משוקלל לפי מספר העובדים, ולכן הוא נמוך ב-1.5% בערך.
    כאן הסדרה מחושבת מחדש, משוקללת כראוי.

    האוכלוסייה בלוח היא "כלל העובדים" — כוללת עצמאים, ולכן רחבה מזו שבשאר
    הדשבורד (שכירים בלבד). הסדרה מוצגת בנפרד ואינה מעורבבת עם נתוני הלמ"ס.
    """
    wb = openpyxl.load_workbook(SRC_BTL, data_only=True)
    ws = wb["8"]
    years = [ws.cell(4, 3 + i).value for i in range(7)]
    if [str(y) for y in years] != [str(y) for y in range(2016, 2023)]:
        sys.exit(f"שנות הלוח אינן 2016–2022: {years}")

    def find(label):
        for r in range(5, ws.max_row + 1):
            if str(ws.cell(r, 2).value or "").strip() == label:
                return r
        sys.exit(f"לא נמצאה שורת {label!r} בלוח")

    def series(r):
        return ([float(ws.cell(r, 3 + i).value) for i in range(7)],
                [float(ws.cell(r, 10 + i).value) for i in range(7)])

    nat_row = 5
    if str(ws.cell(nat_row, 1).value or "").strip() != 'סה"כ- נפה ומחוז':
        sys.exit("שורת הסך הארצי אינה במקומה הצפוי")
    nat_w, nat_s = series(nat_row)
    tz_w, tz_s = series(find("צפת"))
    gl_w, gl_s = series(find("גולן"))
    wb.close()

    out = []
    for i, y in enumerate(years):
        workers = tz_w[i] + gl_w[i]
        salary = (tz_w[i] * tz_s[i] + gl_w[i] * gl_s[i]) / workers
        out.append({"year": str(y), "workers": round(workers),
                    "salary": round(salary, 1), "national": round(nat_s[i], 1),
                    "nationalWorkers": round(nat_w[i]),
                    "ratio": round(salary / nat_s[i], 4)})
    return {"rows": out, "population": "כלל העובדים (שכירים ועצמאים)",
            "measure": "שכר ממוצע לחודש עבודה",
            "source": 'הביטוח הלאומי — שכר והכנסה מעבודה לפי יישוב, לוח 8 (מחוז ונפה)',
            "note": "רמת האשכול חושבה מחיבור נפת צפת ונפת גולן, משוקלל לפי מספר העובדים"}


def read_change(names):
    """מחזיר את נתוני 2022 לכל רשות: מספר שכירים ושכר חודשי ממוצע."""
    wb = openpyxl.load_workbook(SRC_2022, data_only=True)
    ws = wb["רשויות_ענף_כלכלי__סדר_"]
    if (ws.cell(1, 3).value or "").strip() != "מספר שכירים":
        sys.exit("מבנה קובץ 2022 השתנה — עמודה 3 אינה 'מספר שכירים'")

    per, cells, floor = {}, 0, None
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        workers, salary = ws.cell(r, 3).value, ws.cell(r, 4).value
        if not isinstance(name, str) or not name.strip() or workers is None or salary is None:
            continue
        workers, salary = float(workers), float(salary)
        per.setdefault(name.strip(), []).append((workers, salary))
        cells += 1
        floor = workers if floor is None else min(floor, workers)
    wb.close()

    if floor is None or floor < SUPPRESSION_FLOOR:
        sys.exit(f"סף ההשמטה בקובץ 2022 אינו {SUPPRESSION_FLOOR} כמצופה (נמצא {floor})")
    if set(per) != set(names):
        sys.exit(f"רשימת הרשויות ב-2022 אינה תואמת: "
                 f"חסרות {set(names) - set(per)} · עודפות {set(per) - set(names)}")

    out = {}
    for name, rows in per.items():
        w = sum(x for x, _ in rows)
        out[name] = {"workers": round(w),
                     "salary": round(sum(x * y for x, y in rows) / w, 2),
                     "sections": len(rows)}
    # חסם עליון על ההשמטה: כל תא חסר נושא לכל היותר 9 שכירים
    missing = sum(20 - v["sections"] for v in out.values())   # 19 סדרים + "ללא ענף"
    total = sum(v["workers"] for v in out.values())
    return {"year": "2022", "byAuthority": out, "cells": cells,
            # בקובץ 2022 נמסרו שני מדדים בלבד — אין בו חודשי עבודה ואין נתון ארצי
            "metrics": ["workers", "salary"],
            "floor": SUPPRESSION_FLOOR,
            "maxMissing": missing * (SUPPRESSION_FLOOR - 1), "workers": total,
            "coverage": round(1 - missing * (SUPPRESSION_FLOOR - 1) / total, 4)}


def read_socio():
    """
    מחלץ מקובץ הרשויות המקומיות של הלמ"ס את האשכול החברתי-כלכלי של כל רשות,
    ומחשב את השכר החודשי הממוצע בכל אחד מ-10 האשכולות.

    השנים נקראות מהקובץ עצמו — שורה 3 נושאת שנת עדכון לכל עמודה — ולכן החלפת
    הקובץ במהדורה עדכנית יותר מספיקה, בלי לגעת בקוד. ההגדרה של השכר זהה לזו שבקובץ
    העיבוד — "סך שכר ברוטו לפרט מכל משרות העבודה במהלך השנה מחולק במספר חודשי
    העבודה" — ולכן המדדים ברי-השוואה, אך השנים אינן זהות. הפער נשמר ומדווח.
    """
    wb = openpyxl.load_workbook(SRC_RASH, data_only=True)
    ws = wb["נתונים פיזיים ונתוני אוכלוסייה "]
    NAME, CLUSTER, COUNT, SALARY = 1, 251, 138, 139
    NATIONAL_ROW, FIRST_AUTHORITY_ROW = 6, 10   # 7–9 הן שורות סיכום לפי סוג רשות
    if (ws.cell(4, COUNT).value or "").strip() != "בעלי הכנסה משכר" \
            or (ws.cell(5, SALARY).value or "").strip() != "שכר ממוצע" \
            or (ws.cell(NATIONAL_ROW, NAME).value or "").strip() != "כלל ארצי":
        sys.exit("מבנה קובץ הרשויות המקומיות השתנה — יש לעדכן את מיפוי העמודות")

    def num(v):
        try:
            return float(str(v).replace(",", ""))
        except (TypeError, ValueError):
            return None

    of_authority, own_salary, totals = {}, {}, {}
    for r in range(FIRST_AUTHORITY_ROW, ws.max_row + 1):
        name = ws.cell(r, NAME).value
        if not isinstance(name, str) or not name.strip():
            continue
        clu, cnt, sal = (num(ws.cell(r, c).value) for c in (CLUSTER, COUNT, SALARY))
        if not (clu and cnt and sal):
            continue
        clu = int(clu)
        of_authority[name.strip()] = clu
        own_salary[name.strip()] = round(sal, 2)
        t = totals.setdefault(clu, [0.0, 0.0, 0])
        t[0] += cnt * sal
        t[1] += cnt
        t[2] += 1

    national = num(ws.cell(NATIONAL_ROW, SALARY).value)
    clusters = {str(k): {"salary": round(v[0] / v[1], 2), "workers": round(v[1]),
                         "authorities": v[2]}
                for k, v in sorted(totals.items())}

    # אימות: השחזור המשוקלל חייב להתלכד עם שורת "כלל ארצי" שבקובץ
    w = sum(v["workers"] for v in clusters.values())
    mine = sum(v["salary"] * v["workers"] for v in clusters.values()) / w
    wb.close()
    if national and abs(mine - national) / national > 0.005:
        sys.exit(f"ממוצע משוחזר {mine:.2f} אינו תואם את שורת כלל ארצי {national:.2f}")

    # שורה 3 בקובץ נושאת את שנת העדכון של כל עמודה
    salary_year = str(ws.cell(3, COUNT).value or "").strip()
    cluster_year = str(ws.cell(3, CLUSTER).value or "").strip()
    if not (salary_year and cluster_year):
        sys.exit("לא נמצאו שנות העדכון בשורה 3 של קובץ הרשויות המקומיות")

    return {"of_authority": of_authority, "ownSalary": own_salary, "clusters": clusters,
            "national": round(national, 2),
            "year": salary_year, "clusterYear": cluster_year,
            "source": f"הלמ\"ס — קובץ רשויות מקומיות; שכר {salary_year}, "
                      f"אשכול חברתי-כלכלי {cluster_year}"}


def read_meta(wb):
    """שולף מגיליון 'מידע נילווה' את פרטי ההפקה ואת ההגדרות — מילה במילה מהמקור."""
    ws = wb["מידע נילווה"]
    rows = [[(c or "").strip() if isinstance(c, str) else c for c in r]
            for r in ws.iter_rows(min_col=1, max_col=2, values_only=True)]

    def field(prefix):
        for a, b in rows:
            if isinstance(a, str) and a.startswith(prefix) and b:
                return str(b).strip()
        return ""

    # ההגדרות הן השורות שאחרי הכותרת "הסבר נוסף"
    defs, seen_header = [], False
    for a, b in rows:
        if isinstance(a, str) and a.startswith("הסבר נוסף"):
            seen_header = True
            continue
        if seen_header and a and b:
            defs.append({"term": str(a).strip(), "text": str(b).strip()})
    if len(defs) != 3:
        sys.exit(f"צפויות 3 הגדרות בגיליון 'מידע נילווה', נמצאו {len(defs)}")

    return {
        "unit": field('יחידה בלמ"ס'),
        "kind": field("סוג ההפקה"),
        "definitions": defs,
        "sheet": "מידע נילווה",
    }


def read_authorities(wb, socio):
    ws = wb["לפי רשויות בנפת צפת וגולן"]
    rows = list(ws.iter_rows(values_only=True))
    of_nafa = {name: nafa for nafa, names in NAFOT.items() for name in names}

    national = rows[2]           # סך כללי (באלפים)
    nafot = []
    for r in rows[3:5]:
        nafot.append({"name": r[1].strip(), "workers": num(r[3]),
                      "months": num(r[4]), "salary": num(r[5])})

    # הנפות משמשות לאימות מול המקור בלבד — הדשבורד מציג את האשכול כיחידה אחת
    items = []
    for r in rows[5:23]:
        name = (r[2] or "").strip()
        if not name:
            continue
        if name not in of_nafa:
            sys.exit(f"רשות לא מוכרת בשיוך לנפה: {name}")
        if name not in socio["of_authority"]:
            sys.exit(f"רשות ללא אשכול חברתי-כלכלי בקובץ הרשויות: {name}")
        clu = socio["of_authority"][name]
        own = socio["ownSalary"][name]
        items.append({"name": name, "nafa": of_nafa[name], "workers": num(r[3]),
                      "months": num(r[4]), "salary": num(r[5]),
                      "socio": clu,
                      # שכר הרשות בשנת הרף — מאפשר השוואה שנה-מול-שנה מול האשכול
                      "socioOwn": own,
                      "socioRatio": round(own / socio["clusters"][str(clu)]["salary"], 4)})

    # אימות: סכום השכירים ברשויות חייב להתאים לסכום הנפות
    for nf in nafot:
        s = sum(i["workers"] for i in items if i["nafa"] == nf["name"])
        assert abs(s - nf["workers"]) < 1, f"{nf['name']}: {s} != {nf['workers']}"

    total_w = sum(i["workers"] for i in items)
    region = {
        "name": "אשכול גליל מזרחי",
        "workers": total_w,
        "salary": round(sum(i["workers"] * i["salary"] for i in items) / total_w, 2),
        "months": round(sum(i["workers"] * i["months"] for i in items) / total_w, 3),
    }
    return {
        "items": items,
        "nafot": nafot,
        "region": region,
        "national": {"name": "ממוצע ארצי", "workers": round(national[3] * 1000),
                     "months": num(national[4]), "salary": num(national[5])},
    }


def read_anafim(wb, dic):
    ws = wb["לפי ענף כלכלי"]
    rows = list(ws.iter_rows(values_only=True))
    items = []
    for r in rows[3:54]:
        code = str(r[1]).strip()
        if not code or code == "None":
            continue
        codes = [c.strip().zfill(2) for c in code.split("+")]
        seders, parts = [], []
        for c in codes:
            d = dic.get(c)
            if not d:
                sys.exit(f"קוד ענף חסר במילון: {c}")
            if d["seder"] and d["seder"] not in seders:
                seders.append(d["seder"])
            parts.append(d["name"])
        if code not in ANAF_LABELS:
            sys.exit(f"חסר שם תצוגה לקבוצת הענפים: {code}")
        items.append({
            "code": code,
            "label": ANAF_LABELS[code],
            "seder": " · ".join(seders),
            "parts": parts,
            "nat": {"workers": num(r[2]), "months": num(r[3]), "salary": num(r[4])},
            "reg": {"workers": num(r[5]), "months": num(r[6]), "salary": num(r[7])},
        })

    known = {i["code"] for i in items}
    for cid, cname, codes in CLUSTERS:
        missing = [c for c in codes if c not in known]
        if missing:
            sys.exit(f"קיבוץ {cid}: קודים שאינם בנתונים {missing}")
    return items


def main():
    dic = read_dictionary()
    wb = openpyxl.load_workbook(SRC_DATA, data_only=True)
    meta = read_meta(wb)
    socio = read_socio()
    authorities = read_authorities(wb, socio)
    change = read_change([i["name"] for i in authorities["items"]])
    btl = read_btl_trend()
    anafim = read_anafim(wb, dic)
    wb.close()

    data_year = "2024"
    data = {
        "meta": {
            "year": data_year,
            "cluster": "אשכול גליל מזרחי",
            "source": "הלשכה המרכזית לסטטיסטיקה — עיבוד מיוחד מתוך קובצי הכנסה מנהליים",
            "processing": "עיבוד: מרכז הידע האזורי גליל מזרחי | מערבי",
            **meta,
        },
        "metrics": [
            {"id": "salary",  "label": "שכר חודשי ממוצע",   "short": "שכר",         "unit": "₪",     "dec": 0},
            {"id": "workers", "label": "מספר שכירים",        "short": "שכירים",      "unit": "",      "dec": 0},
            {"id": "months",  "label": "ממוצע חודשי עבודה",  "short": "חודשי עבודה", "unit": "חודשים", "dec": 2},
        ],
        "authorities": authorities,
        "anafim": anafim,
        "clusters": [{"id": c, "label": l, "codes": codes} for c, l, codes in CLUSTERS],
        "socio": {k: socio[k] for k in
                  ("clusters", "national", "year", "clusterYear", "source")},
        "change": change,
        "btl": btl,
    }

    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.json")
    with open(out, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=1)

    reg_sum = sum(i["reg"]["workers"] for i in anafim if isinstance(i["reg"]["workers"], (int, float)))
    nat_sum = sum(i["nat"]["workers"] for i in anafim if isinstance(i["nat"]["workers"], (int, float)))
    print(f"נכתב {out}")
    print(f"  רשויות: {len(authorities['items'])} · ענפים: {len(anafim)} · קיבוצים: {len(CLUSTERS)}")
    print(f"  כיסוי שכירים לפי ענף — אשכול {reg_sum:,} מתוך {authorities['region']['workers']:,}"
          f" · ארצי {nat_sum:,} מתוך {authorities['national']['workers']:,}")
    print(f"  שינוי {change['year']}→{data_year}: שכירים "
          f"{change['workers']:,} → {authorities['region']['workers']:,}"
          f" · כיסוי 2022 {change['coverage']*100:.1f}% לפחות"
          f" (עד {change['maxMissing']:,} שכירים בתאים שהושמטו)")
    f, l = btl["rows"][0], btl["rows"][-1]
    print(f"  מגמת ביטוח לאומי {f['year']}–{l['year']}: שכר אשכול "
          f"{f['salary']:,.0f} → {l['salary']:,.0f} ₪ · ארצי {f['national']:,.0f} → {l['national']:,.0f} ₪"
          f" · יחס {f['ratio']*100:.1f}% → {l['ratio']*100:.1f}%")
    used = sorted({i["socio"] for i in authorities["items"]})
    print(f"  אשכולות חברתיים-כלכליים בשימוש: {used}"
          f" · ארצי {socio['year']}: {socio['national']:,.0f} ₪"
          f" (יחס 2024/{socio['year']}: {authorities['national']['salary'] / socio['national']:.4f})")
    for d in meta["definitions"]:
        print(f"  הגדרה מהמקור · {d['term']}: {d['text'][:60]}…")


if __name__ == "__main__":
    main()
