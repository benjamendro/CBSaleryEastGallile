"""Independent readers for the source workbooks.

Everything here opens the spreadsheets with openpyxl and locates values by the
layout a person reads on screen — sheet name, header row, authority name, year
column. **Nothing in this module imports build_data.py, build_btl.py or
btl_read.py**, on purpose: a comparison that reused the pipeline's own reader
would only prove the pipeline agrees with itself.

Where the pipeline is documented to transform a value rather than copy it (the
National Insurance cluster row, which the source averages wrongly), the readers
return the raw cells and the test does the documented arithmetic itself.
"""

import os
import re

import openpyxl

from . import paths

# --- CBS 2024: authorities ------------------------------------------------------

AUTH_SHEET = "לפי רשויות בנפת צפת וגולן"
ANAF_SHEET = "לפי ענף כלכלי"
INFO_SHEET = "מידע נילווה"


def _num(value):
    """A cell's numeric value, or None when the cell is a label or blank.

    The workbooks mix numbers and text in the same columns (header rows, footnote
    markers, counts stored as text), so every read goes through this.
    """
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip().replace(",", ""))
    except ValueError:
        return None


def _rows(path, sheet):
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        yield from workbook[sheet].iter_rows(values_only=True)
    finally:
        workbook.close()


def cbs_authorities():
    """Read the authorities sheet into (authorities, sub_districts, national).

    Layout: a header on row 2, the national total on row 3 in *thousands*, the two
    sub-district rows next (column B), then one row per authority (column C).

    Note what is **not** here: the sheet never says which authority sits in which
    sub-district. That mapping lives in the pipeline, so it cannot be checked
    cell-against-cell — the tests verify it through the sums instead.
    """
    authorities, sub_districts, national = {}, {}, None
    for row in _rows(paths.source_workbook("anaf_2024"), AUTH_SHEET):
        label, nafa, authority = row[0], row[1], row[2]
        workers, months, salary = _num(row[3]), _num(row[4]), _num(row[5])
        if workers is None:
            continue
        if label and "סך כללי" in str(label):
            # published in thousands, unlike every other row on the sheet
            national = {
                "workers": round(workers * 1000),
                "months": months,
                "salary": salary,
            }
        elif nafa:
            sub_districts[str(nafa).strip()] = {
                "workers": round(workers),
                "months": months,
                "salary": salary,
            }
        elif authority:
            authorities[str(authority).strip()] = {
                "workers": round(workers),
                "months": months,
                "salary": salary,
            }
    return authorities, sub_districts, national


def cbs_industries():
    """Read the industry sheet into {code: {'nat': {...}, 'reg': {...}}}.

    Columns C–E are the national reading and F–H the two sub-districts. The C2
    header says '(באלפים)' and is wrong — these are full counts, which is why the
    reader does not scale them.
    """
    industries = {}
    for row in _rows(paths.source_workbook("anaf_2024"), ANAF_SHEET):
        code = row[1]
        national = _num(row[2])
        if code is None or national is None or str(code).strip() in ("", "ענף כלכלי"):
            continue
        regional = _num(row[5])
        industries[str(code).strip()] = {
            "nat": {"workers": round(national), "months": _num(row[3]), "salary": _num(row[4])},
            # a suppressed regional cell stays None rather than becoming a zero
            "reg": {
                "workers": None if regional is None else round(regional),
                "months": _num(row[6]),
                "salary": _num(row[7]),
            },
        }
    return industries


def cbs_definitions():
    """The three explanatory definitions, as written in the 'הסבר נוסף' block."""
    definitions = {}
    seen_heading = False
    for row in _rows(paths.source_workbook("anaf_2024"), INFO_SHEET):
        term, text = row[0], row[1]
        if term and "הסבר נוסף" in str(term):
            seen_heading = True
            continue
        if seen_heading and term and text:
            definitions[str(term).strip()] = str(text).strip()
    return definitions


def cbs_industry_by_authority():
    """The supplementary industry×authority workbook, as {authority: {code: {...}}}."""
    workbook = openpyxl.load_workbook(
        paths.source_workbook("anaf_by_auth"), data_only=True, read_only=True
    )
    sheet = workbook[workbook.sheetnames[0]]
    cells = {}
    try:
        for row in sheet.iter_rows(values_only=True):
            code, authority = row[0], row[4]
            workers, months, salary = _num(row[1]), _num(row[2]), _num(row[3])
            if authority is None or workers is None or str(code).strip() == "ענף כלכלי":
                continue
            cells.setdefault(str(authority).strip(), {})[str(code).strip()] = {
                "workers": round(workers),
                "months": months,
                "salary": salary,
            }
    finally:
        workbook.close()
    return cells


def cbs_2022_by_seder():
    """The 2022 workbook at industry-order level: {authority: {order: {...}}}.

    CLAUDE.md pitfall 5 — 2022 is summed from the order level, not the two-digit
    level, because the suppression floor loses 5.2% of employees at the fine
    level against at most 0.5% here.
    """
    rows = _rows(paths.source_workbook("anaf_2022"), "רשויות_ענף_כלכלי__סדר_")
    orders = {}
    for row in rows:
        authority, order = row[0], row[1]
        workers, salary = _num(row[2]), _num(row[3])
        if not authority or authority == "ShemMaamadMuniOMoeza" or workers is None:
            continue
        orders.setdefault(str(authority).strip(), {})[str(order).strip()] = {
            "workers": round(workers),
            "salary": salary,
        }
    return orders


def cbs_industry_dictionary():
    """The CBS industry dictionary as {two-digit code: (industry name, order name)}.

    Codes are stored zero-padded ('01'); the processing files use them unpadded
    ('1'), so the key is normalised here.
    """
    workbook = openpyxl.load_workbook(
        paths.source_workbook("dictionary"), data_only=True, read_only=True
    )
    entries = {}
    try:
        rows = workbook["רשימת הערכים במילון ענפי כלכלה"].iter_rows(values_only=True)
        next(rows)  # header
        for row in rows:
            code = row[12]
            if code is None:
                continue
            key = str(code).strip().lstrip("0") or "0"
            entries[key] = (str(row[13]).strip(), str(row[18]).strip())
    finally:
        workbook.close()
    return entries


# --- National Insurance table 8 ------------------------------------------------

BTL_YEARS = [2016, 2017, 2018, 2019, 2020, 2021, 2022]


def btl_table8_sub_districts():
    """The Safed and Golan rows of table 8: {nafa: {year: (workers, salary)}}.

    Columns C–I hold headcounts by year and J–P the wages, per the two header
    rows. The sheet's own 'גליל מזרחי' row is deliberately not read: CLAUDE.md
    pitfall 4 records that its wage column is an unweighted mean of these two.
    """
    wanted = {"צפת", "גולן"}
    found = {}
    for row in _rows(paths.source_workbook("btl_table8"), "8"):
        name = str(row[1]).strip() if row[1] else ""
        if name in wanted and name not in found:
            found[name] = {
                year: (_num(row[2 + index]), _num(row[9 + index]))
                for index, year in enumerate(BTL_YEARS)
            }
    return found


def btl_table8_national():
    """The national row of table 8: {year: (workers, salary)}."""
    for row in _rows(paths.source_workbook("btl_table8"), "8"):
        if row[0] and 'סה"כ' in str(row[0]):
            return {
                year: (_num(row[2 + index]), _num(row[9 + index]))
                for index, year in enumerate(BTL_YEARS)
            }
    raise AssertionError("table 8 has no national total row")


def btl_table8_cluster_row():
    """The workbook's own 'גליל מזרחי' row — the one the pipeline must not copy."""
    for row in _rows(paths.source_workbook("btl_table8"), "מחוזות"):
        # exact match: 'צפון ללא גליל מזרחי' is a different row that contains this name
        if row[0] and str(row[0]).strip() == "גליל מזרחי":
            return {
                year: (_num(row[1 + index]), _num(row[8 + index]))
                for index, year in enumerate(BTL_YEARS)
            }
    raise AssertionError("the 'מחוזות' sheet has no East Galilee row")


def weighted(pairs):
    """Combine (weight, value) pairs the way the project's weighting rule requires."""
    total = sum(weight for weight, _ in pairs)
    return sum(weight * value for weight, value in pairs) / total


# --- National Insurance: the 175 appendix tables --------------------------------

BTL_NEW_YEARS = [2023, 2024]
BTL_OLD_YEARS = [2016, 2017, 2018, 2019, 2020, 2021, 2022]

# (edition, appendix) -> (locality table number, regional-council table number).
# CLAUDE.md pitfall 9: table numbers repeat between editions, so the key is the
# pair, never the number alone.
BTL_TABLES = {
    ("new", "א"): (2, 3), ("new", "ד"): (50, 51), ("new", "ז"): (98, 99),
    ("new", "ח"): (114, 115), ("new", "ג"): (34, 35), ("new", "ו"): (82, 83),
    ("old", "א"): (3, 6), ("old", "ד"): (100, 103), ("old", "ז"): (169, 172),
    ("old", "ח"): (192, 195), ("old", "ג"): (77, 80), ("old", "ו"): (146, 149),
}
BTL_EDITION_TITLE = {"new": "2024-2023", "old": "2022-2016"}

# Column offsets of each measure's first year, by edition.
BTL_COLUMNS = {
    "new": {"workers": 2, "mean_year": 5, "median_year": 8, "mean_work": 11, "median_work": 14},
    "old": {"workers": 1, "mean_work": 8},
}

# The locality and council tables total only their own geography, so for
# 2016–2022 the national figure has to come from the district/sub-district
# tables, which carry their own table numbers and an extra label column.
BTL_NATIONAL_OLD = {"ג": 82, "ו": 151, "א": 8}
BTL_DISTRICT_COLUMNS = {"workers": 2, "mean_work": 9}

_BTL_INDEX = None


def _btl_index(directory):
    """{(table number, edition): filename} built from each workbook's own title."""
    global _BTL_INDEX
    if _BTL_INDEX is not None:
        return _BTL_INDEX
    _BTL_INDEX = {}
    for name in sorted(os.listdir(directory)):
        if not name.endswith(".xlsx"):
            continue
        workbook = openpyxl.load_workbook(os.path.join(directory, name), read_only=True)
        sheet = workbook.sheetnames[0]
        title = str(next(workbook[sheet].iter_rows(max_row=1, values_only=True))[0] or "")
        workbook.close()
        for edition, marker in BTL_EDITION_TITLE.items():
            if marker in title:
                _BTL_INDEX[(sheet.strip(), edition)] = name
    return _BTL_INDEX


def btl_rows(directory, number, edition, geo):
    """{authority: row} for one appendix table, totals only.

    The name column is written once per authority and left blank on the gender
    rows beneath it, so it is carried forward. In the 2023–2024 edition only the
    'סה"כ' gender row is kept; the older tables have no gender split at all.
    """
    key = (str(number), edition)
    index = _btl_index(directory)
    if key not in index:
        raise AssertionError(f"no table {number} in edition {edition} among the workbooks")
    workbook = openpyxl.load_workbook(
        os.path.join(directory, index[key]), data_only=True, read_only=True
    )
    rows = {}
    current = None
    try:
        for row in workbook[workbook.sheetnames[0]].iter_rows(values_only=True):
            if row[0]:
                # Table 45 carries a line break inside 'מבואות החרמון'; a reader that
                # does not collapse whitespace loses that council from that table alone.
                current = re.sub(r"\s+", " ", str(row[0])).strip()
            if current is None:
                continue
            if edition == "new":
                if not row[1] or str(row[1]).strip() != 'סה"כ':
                    continue
            rows.setdefault(current, row)
    finally:
        workbook.close()
    return rows


def btl_value(row, edition, measure, year):
    """One cell of a table row, addressed by measure and year."""
    years = BTL_NEW_YEARS if edition == "new" else BTL_OLD_YEARS
    if year not in years or measure not in BTL_COLUMNS[edition]:
        return None
    return _num(row[BTL_COLUMNS[edition][measure] + years.index(year)])


def btl_table_for(directory, appendix, edition, is_council):
    """The right table for an appendix, edition and geography level."""
    locality, council = BTL_TABLES[(edition, appendix)]
    return btl_rows(directory, council if is_council else locality, edition, "rc" if is_council else "loc")


def btl_national_wage(directory, appendix, year):
    """The country-wide wage per month worked for one appendix and year.

    For 2023–2024 the locality table's own 'סה"כ' row is the national total. For
    2016–2022 it is not — that row totals only localities of 2,000+ residents —
    so the figure is taken from the district/sub-district table instead.
    """
    if year in BTL_NEW_YEARS:
        rows = btl_table_for(directory, appendix, "new", False)
        total = next((row for name, row in rows.items() if name.startswith('סה"כ')), None)
        return None if total is None else btl_value(total, "new", "mean_work", year)

    rows = btl_rows(directory, BTL_NATIONAL_OLD[appendix], "old", "district")
    total = next((row for name, row in rows.items() if 'סה"כ' in name), None)
    if total is None:
        return None
    column = BTL_DISTRICT_COLUMNS["mean_work"] + BTL_OLD_YEARS.index(year)
    return _num(total[column])


# --- the income-group tables (the wage distribution) ----------------------------

# (population appendix) -> (locality table, regional-council table), 2023–2024 only.
BTL_INCOME_TABLES = {"ג": (43, 45), "ו": (91, 93), "א": (11, 13)}

# Seven bands, each published as a 2023 column followed by a 2024 column.
BTL_INCOME_FIRST_COLUMN = 6
BTL_INCOME_BANDS = 7


def btl_income_rows(directory, appendix, is_council):
    locality, council = BTL_INCOME_TABLES[appendix]
    return btl_rows(directory, council if is_council else locality, "new", "rc" if is_council else "loc")


def btl_income_bands(row, year):
    """The seven shares of a row, in the order the table prints them."""
    offset = BTL_NEW_YEARS.index(year)
    return [
        _num(row[BTL_INCOME_FIRST_COLUMN + 2 * band + offset])
        for band in range(BTL_INCOME_BANDS)
    ]
