# -*- coding: utf-8 -*-
"""קוראים משותפים ללוחות הביטוח הלאומי, וניתוב רשויות האשכול.

מרכז את הפרטים שנתפסו באימות ואסור לחזור עליהם:

  · **מפתח הלוח הוא (מספר, מהדורה)** — מספרי הלוחות חוזרים בין המהדורות.
    „לוח 77” הוא גם שכר לפי יישוב 2016–2022 וגם קבוצת הכנסה של עצמאים 2023–2024.
  · **ניתוב יישוב מול מועצה אזורית** לפי „צורת יישוב” שבקובץ הצימודים — ובשתי
    האיותים (שם רשימת האשכול והשם הרשמי), אחרת „גליל עליון” לא ימצא את
    „הגליל העליון” וינותב ללוח היישובים שבו הוא אינו קיים.
  · **האיות של הביטוח הלאומי** לשני שמות, שאושר ידנית: „קריית שמונה”, „טובא-זנגרייה”.
  · **„כלל העובדים” לעולם מחושב כ-ד+ז+ח ולא נקרא מהלוח** — עמודת 2019 בלוח 3 של
    מהדורת 2016–2022 מעורבבת בין שורות ב-49 יישובים.

שימוש:
    export BTL_DIR=/path/to/relevant_tables      # תיקיית ה-xlsx שחולצו מה-rar
    from btl_read import load_tables, read, find, geo_of_auth, APPENDIX
"""
import difflib
import os
import sys
import warnings

import openpyxl

warnings.filterwarnings("ignore")

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(REPO, "eshkol-matching"))
from eshkol_matcher import sanitize  # noqa: E402

import btl_catalog  # noqa: E402  — סיווג הלוחות חי במקום אחד בלבד

BTL_DIR = os.environ.get("BTL_DIR", os.path.join(REPO, "relevant_tables"))
TOTAL = 'סה"כ'

# איות הביטוח הלאומי לשמות שנבדלים מקובץ הצימודים. אושר 30/08/2026.
# נשמר כאן ולא כ-fuzzy בזמן ריצה — ראו eshkol-matching/SKILL.md §1.
BTL_SPELLING = {
    "קרית שמונה": "קריית שמונה",
    "טובא-זנגריה": "טובא-זנגרייה",
}

# מבנה החלוקה. אומת לאפס סטייה: א׳ = ד׳+ז׳+ח׳ · ג׳ = ד׳+ח׳ · ו׳ = ז׳+ח׳.
APPENDIX = {
    "א": "כלל העובדים",
    "ב": "שכירים · הכנסה שכירה ועצמאית",
    "ג": "שכירים · הכנסה שכירה בלבד",
    "ד": "שכירים שאינם עצמאים",
    "ה": "עצמאים · הכנסה עצמאית ושכירה",
    "ו": "עצמאים · הכנסה עצמאית בלבד",
    "ז": "עצמאים שאינם שכירים",
    "ח": "שכירים וגם עצמאים",
}
PARTS = ("ד", "ז", "ח")  # שלושת אבני הבניין הזרות זו לזו


# ---------------------------------------------------------------- הרשויות
def load_targets(cluster="גליל מזרחי"):
    """רשויות האשכול מקובץ הצימודים, עם כל האיותים ועם צורת היישוב."""
    path = os.path.join(REPO, "eshkol-matching", "eshkol_mapping.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True)
    rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    wb.close()
    idx = {h: i for i, h in enumerate(rows[0])}
    out = []
    for r in rows[1:]:
        if r[idx["אשכול"]] == cluster and r[idx["סוג ישות"]] == "רשות":
            name = r[idx["שם ברשימת אשכול"]]
            official = r[idx["שם רשמי בלמס"]]
            out.append({
                "name": name,
                "official": official,
                "form": r[idx["צורת יישוב"]],
                "code": r[idx["קוד למס (סמל)"]],
                # כל האיותים המוכרים, בסדר עדיפות
                "aliases": [a for a in (BTL_SPELLING.get(name), BTL_SPELLING.get(official),
                                        official, name) if a],
            })
    return out


TARGETS = load_targets()
# צורת היישוב היא המפתח לניתוב. חובה לקלוט את שתי האיותים.
_RC_NAMES = {a for t in TARGETS if t["form"] == "מועצה אזורית"
             for a in (t["name"], t["official"])}


def geo_of_auth(name):
    """'rc' ללוח המועצות האזוריות, 'loc' ללוח היישובים."""
    return "rc" if name in _RC_NAMES else "loc"


def find(name, table):
    """תווית הרשות בתוך לוח, לפי סדר הפעולות של SKILL.md §2."""
    idx = {sanitize(k): k for k in table}
    cands = [BTL_SPELLING.get(name), name]
    for t in TARGETS:
        if name in (t["name"], t["official"]) or name in t["aliases"]:
            cands += t["aliases"]
    for c in cands:
        if c and sanitize(c) in idx:
            return idx[sanitize(c)]
    for c in cands:
        if not c:
            continue
        near = difflib.get_close_matches(sanitize(c), list(idx), n=1,
                                         cutoff=btl_catalog.FUZZY_CUTOFF)
        if near:
            return idx[near[0]]
    return None


# ---------------------------------------------------------------- הלוחות
_INDEX = None


def _index():
    """(מספר לוח, מהדורה, רמה גאוגרפית) -> שם קובץ."""
    global _INDEX
    if _INDEX is None:
        _INDEX = {}
        for fname in sorted(os.listdir(BTL_DIR)):
            if not fname.endswith(".xlsx"):
                continue
            path = os.path.join(BTL_DIR, fname)
            wb = openpyxl.load_workbook(path, read_only=True)
            sheet = wb.sheetnames[0]
            title = str(next(wb[sheet].iter_rows(max_row=1, values_only=True))[0] or "")
            wb.close()
            name = title or fname[:-5]
            _INDEX[(sheet.strip(), btl_catalog.edition_of(name),
                    btl_catalog.geo_of(name)[0])] = fname
    return _INDEX


def table_file(number, edition, geo):
    key = (str(number), edition, geo)
    try:
        return _index()[key]
    except KeyError:
        raise KeyError(
            "לא נמצא לוח %s במהדורה %s ברמת %s. המפתח הוא (מספר, מהדורה, רמה) — "
            "מספרי לוחות חוזרים בין המהדורות." % key)


def read(number, edition, geo, cols, skip=5, gendered=True, gender_col=1):
    """קורא לוח לפי (מספר, מהדורה, רמה).

    cols: {שם -> אינדקס עמודה}. מחזיר {תווית: {מין: {שם: float|None}}} כשיש פילוח
    מגדרי, אחרת {תווית: {שם: float|None}}. תאים מושמטים ('.', '..', ריק) -> None.
    התווית מתמלאת קדימה, כי הלוחות כותבים אותה רק בשורה הראשונה של כל רשות.

    **לוחות „מחוז ונפה” בנויים אחרת** — עמודת תווית נוספת (מחוז + נפה), ולכן
    המגדר יושב בעמודה 2 ולא 1 והנתונים מוסטים ב-1. לכן `gender_col`, ולכן
    מפות העמודות שלהם מוגדרות בנפרד ואינן משותפות עם לוחות היישוב/המועצה.
    """
    wb = openpyxl.load_workbook(os.path.join(BTL_DIR, table_file(number, edition, geo)),
                                read_only=True)
    ws = wb[wb.sheetnames[0]]
    out, current = {}, None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < skip:
            continue
        if row[0] is not None and str(row[0]).strip():
            current = str(row[0]).strip()
        if current is None:
            continue
        if gendered:
            if len(row) <= gender_col or row[gender_col] is None:
                continue
            bucket = out.setdefault(current, {}).setdefault(str(row[gender_col]).strip(), {})
        else:
            bucket = out.setdefault(current, {})
        for key, col in cols.items():
            value = row[col] if col < len(row) else None
            try:
                bucket[key] = float(str(value).strip())
            except (TypeError, ValueError):
                bucket[key] = None       # '.' / '..' / ריק = הושמט או אין מקרים
    wb.close()
    return out
