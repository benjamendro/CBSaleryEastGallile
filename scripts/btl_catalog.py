# -*- coding: utf-8 -*-
"""קטלוג לוחות הביטוח הלאומי + מדידת כיסוי של 18 רשויות אשכול גליל מזרחי.

קלט:  תיקייה עם קובצי ה-xlsx שחולצו מ-relevant_tables.rar
פלט:  docs/btl-table-catalog.csv  — שורה לכל לוח: מהדורה, נספח, משפחת מדדים,
      רמה גאוגרפית, פילוח מגדרי, וכמה מרשויות האשכול נמצאו בו בפועל.

הרצה:
    python3 scripts/btl_catalog.py <תיקיית-הלוחות> [נתיב-פלט]

הכלל מ-eshkol-matching/SKILL.md נשמר: התאמה מדויקת → ניקוי → מיפוי ידני →
fuzzy בסף 0.85 לכל הפחות. כל התאמת fuzzy מדווחת בעמודה נפרדת לאישור אנושי.
"""
import csv
import difflib
import json
import os
import re
import sys
import warnings

import openpyxl

warnings.filterwarnings("ignore")

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(REPO, "eshkol-matching"))
from eshkol_matcher import sanitize  # noqa: E402

FUZZY_CUTOFF = 0.85  # SKILL.md §2.4 — לעולם לא מתחת לזה

# ---------------------------------------------------------------- סיווג
# נספחי הביטוח הלאומי. הסדר קובע: הביטוי הספציפי ביותר נבדק ראשון.
POP_RULES = [
    ("ח", "שכיר וגם עצמאי", [r"עובדים שהינם שכירים וגם עצמאים", r"שכירים שהינם גם עצמאים"]),
    ("ד", "שכירים שאינם עצמאים", [r"שכירים שאינם גם עצמאים", r"שכירים שאינם עצמאים"]),
    ("ז", "עצמאים שאינם שכירים", [r"עצמאים שאינם גם שכירים", r"עצמאים שאינם שכירים"]),
    ("ו", "עצמאים · עצמאית בלבד", [r"עצמאים \(הכנסה מעבודה עצמאית בלבד", r"עצמאים \(הכנסה מעבודה עצמאית\)"]),
    ("ה", "עצמאים · עצמאית ושכירה", [r"עצמאים \(הכנסה מעבודה עצמאית ושכירה", r"עצמאים \(הכנסה מעבודה שכירה ועצמאית",
                                     r"אחוז העצמאים ששכרם.*שכירה ועצמאית"]),
    ("ג", "שכירים · שכירה בלבד", [r"שכירים \(הכנסה מעבודה שכירה בלבד", r"ה?שכירים \(הכנסה מעבודה שכירה\)"]),
    ("ב", "שכירים · שכירה ועצמאית", [r"ה?שכירים \(הכנסה מעבודה שכירה ועצמאית", r"אחוז השכירים ששכרם.*שכירה ועצמאית"]),
    ("א", "כלל העובדים", [r"כלל העובדים", r"כלל השכירים", r"אחוז העובדים ששכרם"]),
]

MEASURE_RULES = [
    ("insured", "סוג המבוטחים", lambda n: n.startswith("סוג המבוטחים")),
    ("avg_med", "שכר ממוצע וחציוני", lambda n: "שכר ממוצע וחציוני" in n),
    ("duration", "משך העבודה", lambda n: "משך העבודה" in n),
    ("distrib", "התפלגות קבוצת הכנסה", lambda n: "קבוצת ההכנסה" in n),
    ("minwage", "שיעור עד שכר מינימום", lambda n: "אינו גבוה משכר המינימום" in n),
    ("avg_only", "שכר ממוצע בלבד", lambda n: "שכר ממוצע לחודש" in n),
]


def population_of(name):
    for code, label, pats in POP_RULES:
        if any(re.search(p, name) for p in pats):
            return code, label
    if "סוג המבוטחים" in name:
        return "—", "כלל המבוטחים"
    return "?", "?"


def measure_of(name):
    for code, label, test in MEASURE_RULES:
        if test(name):
            return code, label
    return "?", "?"


def geo_of(name):
    if re.search(r"לפי מועצה א[יז]זורית|לפי מועצה אזורית|מועצה איזורית ומין", name):
        return "rc", "מועצה אזורית"
    if re.search(r"לפי מחוז", name):
        return "district", "מחוז ונפה"
    if re.search(r"לפי יישוב|יישוב ומין|ולפי יישוב", name) or "יישובים עם לפחות" in name:
        return "loc", "יישוב (2,000+)"
    return "?", "?"


def edition_of(name):
    if "2024-2023" in name:
        return "2023–2024"
    if "2022-2016" in name:
        return "2016–2022"
    m = re.search(r",\s*(201[6-9]|202[0-4])\s*\(", name)
    return m.group(1) if m else "?"


# ---------------------------------------------------------------- רשויות היעד
def load_targets(cluster="גליל מזרחי"):
    """18 רשויות האשכול מקובץ הצימודים, מופרדות ליישובים ולמועצות אזוריות."""
    path = os.path.join(REPO, "eshkol-matching", "eshkol_mapping.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True)
    rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    wb.close()
    idx = {h: i for i, h in enumerate(rows[0])}
    out = []
    for r in rows[1:]:
        if r[idx["אשכול"]] == cluster and r[idx["סוג ישות"]] == "רשות":
            out.append({
                "name": r[idx["שם ברשימת אשכול"]],
                "official": r[idx["שם רשמי בלמס"]],
                "form": r[idx["צורת יישוב"]],
                "code": r[idx["קוד למס (סמל)"]],
            })
    return out


def first_column_labels(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    labels = [str(row[0]).strip() for row in ws.iter_rows(min_col=1, max_col=1, values_only=True)
              if row[0] is not None and str(row[0]).strip()]
    sheet, dims = ws.title, (ws.max_row, ws.max_column)
    wb.close()
    return labels, sheet, dims


def match(target, label_index):
    """מחזיר (תווית-בלוח, שיטה) או (None, None). סדר הפעולות לפי SKILL.md §2."""
    for candidate, method in ((target["official"], "official"), (target["name"], "name")):
        if candidate and sanitize(candidate) in label_index:
            return label_index[sanitize(candidate)], method
    for candidate in (target["official"], target["name"]):
        if not candidate:
            continue
        near = difflib.get_close_matches(sanitize(candidate), list(label_index), n=1, cutoff=FUZZY_CUTOFF)
        if near:
            return label_index[near[0]], "fuzzy"
    return None, None


def build(src_dir, long_name_map=None):
    targets = load_targets()
    localities = [t for t in targets if t["form"] != "מועצה אזורית"]
    councils = [t for t in targets if t["form"] == "מועצה אזורית"]
    long_name_map = long_name_map or {}

    catalog = []
    for fname in sorted(os.listdir(src_dir)):
        if not fname.endswith(".xlsx"):
            continue
        title = long_name_map.get(fname, fname)[:-5]
        labels, sheet, (nrows, ncols) = first_column_labels(os.path.join(src_dir, fname))
        pop_code, pop_label = population_of(title)
        geo_code, geo_label = geo_of(title)
        measure_code, measure_label = measure_of(title)

        row = {
            "לוח": sheet,
            "מהדורה": edition_of(title),
            "נספח": pop_code,
            "אוכלוסייה": pop_label,
            "משפחת מדדים": measure_label,
            "רמה גאוגרפית": geo_label,
            "פילוח מגדרי": "כן" if re.search(r"ומין|מגדר", title) else "לא",
            "שורות": nrows,
            "עמודות": ncols,
            "כיסוי": "",
            "חסרות": "",
            "התאמות fuzzy": "",
            "שם הלוח": title,
        }
        if geo_code in ("loc", "rc"):
            index = {sanitize(x): x for x in labels}
            wanted = localities if geo_code == "loc" else councils
            found, missing, fuzzy = 0, [], []
            for t in wanted:
                label, method = match(t, index)
                if label:
                    found += 1
                    if method == "fuzzy":
                        fuzzy.append(f"{t['name']}→{label}")
                else:
                    missing.append(t["name"])
            row["כיסוי"] = f"{found}/{len(wanted)}"
            row["חסרות"] = " · ".join(missing)
            row["התאמות fuzzy"] = " · ".join(fuzzy)
        catalog.append(row)
    return catalog


def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    src = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else os.path.join(REPO, "docs", "btl-table-catalog.csv")
    lnm_path = os.path.join(src, "_longname_map.json")
    lnm = json.load(open(lnm_path, encoding="utf-8")) if os.path.exists(lnm_path) else {}

    catalog = build(src, lnm)
    if not catalog:
        sys.exit(f"לא נמצאו קובצי xlsx בתיקייה {src}")

    with open(out, "w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(catalog[0].keys()))
        writer.writeheader()
        writer.writerows(catalog)

    unclassified = [r for r in catalog if "?" in (r["נספח"], r["משפחת מדדים"], r["רמה גאוגרפית"])]
    print(f"נכתבו {len(catalog)} לוחות → {out}")
    if unclassified:
        print(f"אזהרה: {len(unclassified)} לוחות לא סווגו במלואם:")
        for r in unclassified:
            print("  ", r["שם הלוח"][:90])


if __name__ == "__main__":
    main()
