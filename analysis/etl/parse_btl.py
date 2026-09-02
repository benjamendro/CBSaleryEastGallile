# -*- coding: utf-8 -*-
"""Re-parse the BTL (National Insurance Institute) wage tables from the raw xlsx.

Why not use the pre-made unified_all.csv:
  (a) in every 'מחוז ונפה' table it wrote the sub-district into the gender column
      and dropped the male/female breakdown altogether;
  (b) it is missing 28 of the 180 source tables (all 'סוג המבוטחים', and every
      per-year locality mean+median table for 2016-2022).

Here the sheet is reconstructed from openpyxl merged-cell ranges, so the header
hierarchy and the merged geography cells are exact rather than forward-filled.
"""
import os, re, zipfile
import numpy as np
import pandas as pd
from openpyxl import load_workbook

D = '/d/relevant_tables'
GEO_WORDS = ('מחוז', 'נפה', 'יישוב', 'מועצה', 'צורת')
SEX_WORDS = ('מין', 'מגדר')
SEX_VALUES = {'סה"כ', 'גברים', 'נשים', 'זכר', 'נקבה', 'סך הכל'}


def readable(path):
    try:
        return zipfile.ZipFile(path).testzip() is None
    except Exception:
        return False


def grid_with_merges(path):
    """Return (raw grid, grid with every merged range expanded).

    The raw grid is what tells header rows from data rows: a merged label such as
    'מין' spanning the header block would otherwise make the bare year row look
    like data.  The expanded grid is what the header hierarchy and the merged
    geography cells are read from.
    """
    wb = load_workbook(path, data_only=True, read_only=False)
    ws = wb.active
    raw = [[c.value for c in row] for row in ws.iter_rows()]
    exp = [list(r) for r in raw]
    for rng in ws.merged_cells.ranges:
        v = exp[rng.min_row - 1][rng.min_col - 1]
        for r in range(rng.min_row - 1, rng.max_row):
            for c in range(rng.min_col - 1, rng.max_col):
                exp[r][c] = v
    wb.close()
    return raw, exp


def is_num(v):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return False
    if isinstance(v, (int, float, np.integer, np.floating)):
        return True
    s = str(v).replace(',', '').replace('%', '').strip()
    try:
        float(s); return True
    except ValueError:
        return False


def txt(v):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return ''
    return re.sub(r'\s+', ' ', str(v)).strip()


def parse(path):
    raw, exp = grid_with_merges(path)
    if not exp:
        return None
    ncol = max(len(r) for r in exp)
    pad = lambda g: [list(r) + [None] * (ncol - len(r)) for r in g]
    d, draw = pd.DataFrame(pad(exp)), pd.DataFrame(pad(raw))
    keep = [c for c in d.columns if d[c].map(lambda v: txt(v) != '').any()]
    d, draw = d[keep], draw[keep]
    d.columns = draw.columns = range(d.shape[1])
    ncol = d.shape[1]

    # A data row carries a TEXT label in one of the first cells AND numbers to its
    # right.  Header rows have either no numbers (label rows) or no text label
    # (the bare year row: '2023 | 2024 | אחוז שינוי').
    start = None
    for r in range(1, min(len(draw), 15)):
        row = draw.iloc[r]
        has_label = any(txt(row[c]) != '' and not is_num(row[c]) for c in range(min(3, ncol)))
        if has_label and any(is_num(row[c]) for c in range(ncol)):
            start = r
            break
    if start is None or start < 2:
        return None

    hdr = [[txt(v) for v in d.iloc[r].tolist()] for r in range(1, start)]
    body = d.iloc[start:].reset_index(drop=True)

    # --- key (label) columns vs value columns -----------------------------------
    key_cols = []
    for c in range(ncol):
        nn = [v for v in body[c] if txt(v) != '']
        if not nn:
            continue
        if np.mean([is_num(v) for v in nn]) < 0.5:
            key_cols.append(c)
        else:
            break
    if not key_cols:
        key_cols = [0]

    key_names = []
    for c in key_cols:
        label = next((h[c] for h in hdr if h[c]), '')
        vals = set(txt(v) for v in body[c] if txt(v) != '')
        if any(w in label for w in SEX_WORDS) or (vals and vals <= SEX_VALUES):
            key_names.append('gender')
        else:
            key_names.append('geo%d' % (sum(1 for k in key_names if k.startswith('geo')) + 1))

    # inherit down: merges are already expanded, so this only fills genuinely
    # blank cells — and a new value at a higher level resets the levels below it.
    filled, cur = [], [None] * len(key_cols)
    for _, row in body.iterrows():
        for i, c in enumerate(key_cols):
            v = txt(row[c])
            if v:
                cur[i] = v
                for k in range(i + 1, len(cur)):
                    cur[k] = None
        filled.append(list(cur))
    keys = pd.DataFrame(filled, columns=key_names)
    if 'gender' not in keys:
        keys['gender'] = None
    for gcol in ('geo1', 'geo2'):
        if gcol not in keys:
            keys[gcol] = None

    # --- melt the value block ----------------------------------------------------
    recs = []
    for c in range(max(key_cols) + 1, ncol):
        parts, seen = [], set()
        for h in hdr:
            p = h[c]
            if p and p not in seen:
                seen.add(p); parts.append(p)
        year, keep_parts = None, []
        for p in parts:
            m = re.fullmatch(r'(20\d\d)(\.0)?', p)
            if m:
                year = int(m.group(1))
            else:
                keep_parts.append(p)
        metric = ' | '.join(keep_parts)
        if 'אחוז שינוי' in metric:
            continue                       # derived, recomputable
        for i, v in enumerate(body[c]):
            if is_num(v):
                recs.append((keys.geo1[i], keys.geo2[i], keys.gender[i], year, metric,
                             float(str(v).replace(',', '').replace('%', ''))))
    return pd.DataFrame(recs, columns=['geo1', 'geo2', 'gender', 'year', 'metric', 'value'])


def main():
    frames, failed = [], []
    for f in sorted(os.listdir(D)):
        if not f.endswith('.xlsx'):
            continue
        p = os.path.join(D, f)
        if not readable(p):
            failed.append((f, 'corrupt in archive')); continue
        try:
            t = parse(p)
            if t is None or t.empty:
                failed.append((f, 'no data parsed')); continue
            stem = f[:-5]
            t['src'] = stem
            if t['year'].isna().all():          # single-year table: year is in the filename
                m = re.search(r'(?<!\d)(20\d\d)(?!\s*-\s*20\d\d)(?!\d)', stem)
                if m:
                    t['year'] = int(m.group(1))
            frames.append(t)
        except Exception as e:
            failed.append((f, '%s: %s' % (type(e).__name__, e)))
    all_ = pd.concat(frames, ignore_index=True)
    all_.to_pickle('/d/work/btl_reparsed.pkl')
    print('parsed %d files -> %s rows | null years: %d' % (len(frames), format(len(all_), ','),
                                                           all_.year.isna().sum()))
    print('failed %d (all corrupt in archive: %s)' %
          (len(failed), all(e == 'corrupt in archive' for _, e in failed)))


if __name__ == '__main__':
    main()
